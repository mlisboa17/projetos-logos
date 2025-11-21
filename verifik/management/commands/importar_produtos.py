from django.core.management.base import BaseCommand
from verifik.models import ProdutoMae, CodigoBarrasProdutoMae
from openpyxl import load_workbook


class Command(BaseCommand):
    help = 'Importa produtos de um arquivo Excel'

    def add_arguments(self, parser):
        parser.add_argument('arquivo', type=str, help='Caminho do arquivo Excel')

    def handle(self, *args, **options):
        arquivo = options['arquivo']
        
        self.stdout.write(self.style.WARNING(f'Lendo arquivo: {arquivo}'))
        
        # Ler Excel com openpyxl
        wb = load_workbook(arquivo)
        
        # Tentar encontrar a aba com dados
        ws = None
        if 'Posição estoque' in wb.sheetnames:
            ws = wb['Posição estoque']
        elif 'Posicao estoque' in wb.sheetnames:
            ws = wb['Posicao estoque']
        else:
            ws = wb.active
        
        criados = 0
        atualizados = 0
        erros = 0
        
        # Pular primeira linha (cabeçalho) - linha 2 tem o header
        linhas = list(ws.iter_rows(min_row=3, values_only=True))
        total = len(linhas)
        
        self.stdout.write(self.style.SUCCESS(f'Total de produtos no arquivo: {total}'))
        
        for idx, row in enumerate(linhas, start=2):
            try:
                codigo_barras = str(row[0]).strip() if row[0] else ''
                descricao = str(row[1]).strip() if row[1] else ''
                categoria = str(row[2]).strip() if row[2] else 'CERVEJA'
                preco = float(row[3]) if row[3] else 0.0
                
                if not codigo_barras or not descricao:
                    continue
                
                # Extrair marca da descrição (segunda palavra geralmente é a marca)
                palavras = descricao.split()
                marca = palavras[1] if len(palavras) > 1 else ''
                
                # Verificar se código de barras já existe
                codigo_existente = CodigoBarrasProdutoMae.objects.filter(codigo=codigo_barras).first()
                
                if codigo_existente:
                    # Produto já existe, apenas atualizar preço se necessário
                    produto = codigo_existente.produto_mae
                    if produto.preco != preco:
                        produto.preco = preco
                        produto.save()
                        atualizados += 1
                        self.stdout.write(f'  ✓ Atualizado: {descricao[:50]}... (R$ {preco})')
                    continue
                
                # Criar novo produto
                produto = ProdutoMae.objects.create(
                    descricao_produto=descricao,
                    marca=marca,
                    tipo=categoria,
                    preco=preco,
                    ativo=True
                )
                
                # Criar código de barras
                CodigoBarrasProdutoMae.objects.create(
                    produto_mae=produto,
                    codigo=codigo_barras,
                    principal=True
                )
                
                criados += 1
                self.stdout.write(self.style.SUCCESS(f'  ✓ Criado: {descricao[:50]}... (R$ {preco})'))
                
            except Exception as e:
                erros += 1
                self.stdout.write(self.style.ERROR(f'  ✗ Erro na linha {idx}: {str(e)}'))
        
        # Resumo
        self.stdout.write(self.style.SUCCESS(f'\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━'))
        self.stdout.write(self.style.SUCCESS(f'RESUMO DA IMPORTAÇÃO:'))
        self.stdout.write(self.style.SUCCESS(f'  • Total no arquivo: {total}'))
        self.stdout.write(self.style.SUCCESS(f'  • Criados: {criados}'))
        self.stdout.write(self.style.WARNING(f'  • Atualizados: {atualizados}'))
        if erros > 0:
            self.stdout.write(self.style.ERROR(f'  • Erros: {erros}'))
        self.stdout.write(self.style.SUCCESS(f'━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n'))
